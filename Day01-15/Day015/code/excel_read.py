"""
读取Excel文件

@Author:jyang
@Date:5/22/2019
"""
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook('../res/studentData.xlsx')
print(wb.sheetnames)
sheet = wb[wb.sheetnames[0]]
print(sheet.title)
for row in range(2, 4):
    for col in range(4):
        cell_index = chr(65+col) + str(row)
        print(sheet[cell_index].value, end='\t')
    print()
